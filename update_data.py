import openpyxl
import json
import re

def to_id(name):
    s = str(name).lower()
    s = re.sub(r'[àáạảãâầấậẩẫăằắặẳẵ]', 'a', s)
    s = re.sub(r'[èéẹẻẽêềếệểễ]', 'e', s)
    s = re.sub(r'[ìíịỉĩ]', 'i', s)
    s = re.sub(r'[òóọỏõôồốộổỗơờớợởỡ]', 'o', s)
    s = re.sub(r'[ùúụủũưừứựửữ]', 'u', s)
    s = re.sub(r'[ỳýỵỷỹ]', 'y', s)
    s = re.sub(r'đ', 'd', s)
    s = re.sub(r'[^a-z0-9\s-]', '', s)
    s = re.sub(r'\s+', '-', s).strip('-')
    return s

wb = openpyxl.load_workbook('website - alma dungduong.xlsx', data_only=True)
ws = wb['2.Sản phẩm']

unsplash_imgs = [
    "https://images.unsplash.com/photo-1556228720-195a672e8ff5?q=80&w=800",
    "https://images.unsplash.com/photo-1601049541289-9b1b7bbbfe19?q=80&w=800",
    "https://images.unsplash.com/photo-1612817288484-6f916006741a?q=80&w=800",
    "https://images.unsplash.com/photo-1556228444-7164923f1489?q=80&w=800",
    "https://images.unsplash.com/photo-1556228444-2457636e74fc?q=80&w=800",
    "https://images.unsplash.com/photo-1612817288484-6f916006741a?q=80&w=800",
    "https://images.unsplash.com/photo-1596755389378-c31d21fd1273?q=80&w=800",
    "https://images.unsplash.com/photo-1620916566398-39f1143ab7be?q=80&w=800",
    "https://images.unsplash.com/photo-1556228366-2457636e74fc?q=80&w=800"
]

products = []
for r in range(3, 48):
    name = ws.cell(r, 2).value
    if not name or not str(name).strip():
        continue
    name = str(name).strip()
    
    cat = ws.cell(r, 1).value
    eng = ws.cell(r, 3).value
    flag = ws.cell(r, 4).value
    price = ws.cell(r, 5).value
    sale = ws.cell(r, 6).value
    pct = ws.cell(r, 7).value
    saving = ws.cell(r, 8).value
    vol = ws.cell(r, 9).value
    gift = ws.cell(r, 10).value
    desc = ws.cell(r, 11).value
    ingr = ws.cell(r, 12).value
    cert = ws.cell(r, 13).value
    usage = ws.cell(r, 14).value
    
    pid = to_id(name)
    # Give some specific IDs for known combos and products to keep them stable
    if r == 3: pid = "xit-duong-chuyen-sau-miracle"
    elif r == 4: pid = "tinh-chat-tai-sinh-2-0"
    elif r == 5: pid = "tinh-chat-tai-sinh-2-7"
    elif r == 6: pid = "sua-rua-mat-nuoc-bang-glacier"
    elif r == 7: pid = "kem-chong-nang-smart-suncare"
    elif r == 8: pid = "combo-1-trang-sang"
    elif r == 9: pid = "combo-2-phuc-hoi"
    elif r == 10: pid = "combo-3-toan-dien"
    elif r == 11: pid = "combo-4-tinh-gon"
    
    if r==10:
        cat = "Khuyến mãi" # keeping existing category
        
    p = {
        "id": pid,
        "title": name,
        "category": str(cat).strip() if cat else "Khác",
        "price": int(sale) if sale else int(price),
        "rating": 4.9,
        "reviewsCount": 100 + r,
        "image": unsplash_imgs[(r-3) % len(unsplash_imgs)],
        "features": [],
        "skinConcerns": []
    }
    
    if sale and price:
        p["originalPrice"] = int(price)
        
    if eng: p["englishName"] = str(eng).strip().replace('\n','')
    if flag: p["flag"] = str(flag).strip()
    if vol: p["volume"] = str(vol).strip()
    if gift: p["gift"] = str(gift).strip()
    if desc: p["fullDescription"] = str(desc).strip()
    if ingr: p["ingredients"] = str(ingr).strip()
    if cert: p["certifications"] = str(cert).strip()
    if usage: p["usage"] = str(usage).strip()
    
    # short description
    if desc:
        lines = str(desc).strip().split('\n')
        p["description"] = lines[0][:150] + ("..." if len(lines[0]) > 150 else "")
    else:
        p["description"] = f"{name} - Sản phẩm chăm sóc an toàn, hiệu quả."
        
    products.append(p)

out = "export const products: Product[] = " + json.dumps(products, indent=2, ensure_ascii=False) + ";"
with open('products_out.ts', 'w') as f:
    f.write(out)
print("done")
